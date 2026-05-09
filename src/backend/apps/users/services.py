import datetime
from decimal import Decimal

from django.contrib.auth.hashers import make_password
from django.core.exceptions import ValidationError
from django.db import transaction
from django.shortcuts import get_object_or_404

from apps.departments.models import Bolum
from apps.users.models import Akademisyen, Ogrenci, User

HARF_NOTU_KATSAYI = {
    "AA": Decimal("4.0"),
    "BA": Decimal("3.5"),
    "BB": Decimal("3.0"),
    "CB": Decimal("2.5"),
    "CC": Decimal("2.0"),
    "DC": Decimal("1.5"),
    "DD": Decimal("1.0"),
    "FF": Decimal("0.0"),
}


class UsersService:

    @staticmethod
    def ogr_no_uret():
        yil = datetime.datetime.now().year
        son_ogrenci = (
            Ogrenci.objects.filter(ogr_no__startswith=str(yil))
            .order_by("-ogr_no")
            .first()
        )
        if son_ogrenci:
            son_sira = int(son_ogrenci.ogr_no[4:]) + 1
        else:
            son_sira = 1
        return f"{yil}{son_sira:04d}"

    @staticmethod
    def kullanici_adi_uret(ad, soyad, rol_prefix=""):
        temel = f"{rol_prefix}{ad.lower()}.{soyad.lower()}".replace(" ", "")
        if not User.objects.filter(username=temel).exists():
            return temel
        sayac = 1
        while User.objects.filter(username=f"{temel}{sayac}").exists():
            sayac += 1
        return f"{temel}{sayac}"

    @staticmethod
    def sifre_uret(uzunluk=8):
        import secrets
        import string
        karakterler = string.ascii_letters + string.digits
        return ''.join(secrets.choice(karakterler) for _ in range(uzunluk))

    @staticmethod
    @transaction.atomic
    def ogrenci_olustur(ad, soyad, bolum_kodu, sinif=1, sifre=None):
        bolum = get_object_or_404(Bolum, bolum_kodu=bolum_kodu)
        ogr_no = UsersService.ogr_no_uret()
        kullanici_adi = UsersService.kullanici_adi_uret(ad, soyad)
        sifre = sifre or UsersService.sifre_uret()

        user = User(
            username=kullanici_adi,
            ad=ad,
            soyad=soyad,
            role=User.Role.OGRENCI,
            password=make_password(sifre),
        )
        user.full_clean()
        user.save()

        ogrenci = Ogrenci(
            user=user,
            ogr_no=ogr_no,
            bolum=bolum,
            sinif=sinif,
        )
        ogrenci.full_clean()
        ogrenci.save()
        return ogrenci

    @staticmethod
    def ogrenci_getir(ogrenci_id):
        return get_object_or_404(Ogrenci, pk=ogrenci_id)

    @staticmethod
    def akademisyen_getir(akademisyen_id):
        return get_object_or_404(Akademisyen, pk=akademisyen_id)

    @staticmethod
    def gpa_guncelle(ogrenci):
        from apps.enrollments.models import DersKaydi

        kayitlar = DersKaydi.objects.filter(
            ogrenci=ogrenci,
            onay_durumu=DersKaydi.Durum.ONAYLANDI,
            harf_notu__isnull=False,
        ).select_related("donem_dersi__ders")

        toplam_kredi = Decimal("0")
        agirlikli_toplam = Decimal("0")

        for kayit in kayitlar:
            kredi = Decimal(kayit.donem_dersi.ders.kredi)
            katsayi = HARF_NOTU_KATSAYI.get(kayit.harf_notu, Decimal("0"))
            agirlikli_toplam += kredi * katsayi
            toplam_kredi += kredi

        yeni_gpa = (agirlikli_toplam / toplam_kredi).quantize(Decimal("0.01")) if toplam_kredi > 0 else Decimal("0.00")
        ogrenci.gpa = yeni_gpa
        ogrenci.save(update_fields=["gpa"])
        return ogrenci.gpa

    @staticmethod
    @transaction.atomic
    def ogrenci_kaydi_excel(dosya):
        try:
            import openpyxl
        except ImportError:
            raise ValidationError("openpyxl kurulu değil.")

        try:
            calisma_kitabi = openpyxl.load_workbook(dosya)
        except Exception:
            raise ValidationError("Excel dosyası okunamadı.")

        sayfa = calisma_kitabi.active

        # Şifre kolonu ekle
        if sayfa.cell(1, sayfa.max_column).value != 'Şifre':
            sayfa.cell(1, sayfa.max_column + 1).value = 'Şifre'

        sifre_kolon = sayfa.max_column
        basarili = 0
        hatali = 0
        hatalar = []
        sifreler = {}  # Şifreleri geçici tut

        for satir_no, satir in enumerate(sayfa.iter_rows(min_row=2), start=2):
            if not any(hucre.value for hucre in satir):
                continue
            try:
                ad, soyad, bolum_kodu = satir[0].value, satir[1].value, satir[2].value
                if not all([ad, soyad, bolum_kodu]):
                    raise ValueError("Eksik alan")

                sifre = UsersService.sifre_uret()

                UsersService.ogrenci_olustur(
                    ad=str(ad).strip(),
                    soyad=str(soyad).strip(),
                    bolum_kodu=str(bolum_kodu).strip(),
                    sifre=sifre
                )

                sifreler[satir_no] = sifre  # Şifreyi kaydet
                basarili += 1
            except Exception as hata:
                hatali += 1
                hatalar.append({"satir": satir_no, "hata": str(hata)})

        # HATA VARSA ROLLBACK
        if hatali > 0:
            raise ValidationError(
                f"{hatali} satırda hata var, hiçbir öğrenci kaydedilmedi.",
                params={'hatalar': hatalar}  # Hataları parametre olarak gönder
            )

        # HATA YOKSA şifreleri Excel'e yaz
        for satir_no, sifre in sifreler.items():
            sayfa.cell(satir_no, sifre_kolon).value = sifre

        from io import BytesIO
        output = BytesIO()
        calisma_kitabi.save(output)
        output.seek(0)

        return {
            "basarili": basarili,
            "hatali": hatali,
            "hatalar": hatalar,
            "dosya": output
        }
